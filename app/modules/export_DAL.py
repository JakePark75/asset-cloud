"""
app/modules/export_DAL.py
자산 현황 Excel Export — 3시트 구성
  Sheet1: 계좌별 보유종목
  Sheet2: 보유종목 통합
  Sheet3: 실적기록
"""
import io
import datetime
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.db import get_db, is_us_market
from common.redis_store import get_all_prices, get_redis
from app.modules.history_DAL import load_history, load_today_row

KST = ZoneInfo("Asia/Seoul")

# ── 스타일 상수 ───────────────────────────────────────────────────────────────

_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL  = PatternFill("solid", fgColor="2C3E50")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

_SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
_SECTION_FILL = PatternFill("solid", fgColor="1A252F")

_TOTAL_FONT   = Font(bold=True, size=10)

_META_FONT    = Font(italic=True, color="888888", size=9)


# ── 공통 유틸 ─────────────────────────────────────────────────────────────────

def _usd_krw_from_redis(prices: dict) -> float:
    if "USDKRW=X" in prices:
        return float(prices["USDKRW=X"]["price"])
    try:
        r = get_redis()
        if r:
            raw = r.get("usd_krw")
            if raw:
                return float(raw)
    except Exception:
        pass
    return 1350.0


def _f(v, d=0):
    """float 변환 + 반올림. 변환 실패 시 None."""
    if v is None:
        return None
    try:
        return round(float(v), d)
    except Exception:
        return None


def _write_header_row(ws, row: int, cols: list):
    for col, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font  = _HEADER_FONT
        c.fill  = _HEADER_FILL
        c.alignment = _HEADER_ALIGN


def _auto_width(ws, min_w=8, max_w=28):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0
        )
        ws.column_dimensions[letter].width = min(max_w, max(min_w, max_len + 2))


def _eval_krw(ticker: str, qty_f: float, price: float, market: str, usd_krw: float) -> float:
    if ticker == "KRW":
        return qty_f
    if ticker == "USD":
        return qty_f * usd_krw
    if is_us_market(market):
        return qty_f * price * usd_krw
    return qty_f * price


# ── Sheet 1: 계좌별 보유종목 ──────────────────────────────────────────────────

_S1_COLS = ["종목명", "티커", "시장", "레버리지", "수량", "평단", "현재가", "평가액(원)", "손익(원)", "손익률(%)"]
_S1_NCOLS = len(_S1_COLS)


def _build_sheet1(wb, prices: dict, usd_krw: float, now_str: str):
    ws = wb.active
    ws.title = "계좌별 보유종목"

    ws.cell(1, 1, f"기준: {now_str}").font = _META_FONT
    ws.append([""] * _S1_NCOLS)

    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, name, alias
            FROM accounts
            WHERE is_watch = false OR is_watch IS NULL
            ORDER BY id
        """)
        accounts = cur.fetchall()

        for acct_id, acct_name, acct_alias in accounts:
            label = acct_alias or acct_name

            # 섹션 헤더
            ws.append([label] + [""] * (_S1_NCOLS - 1))
            sec_row = ws.max_row
            ws.merge_cells(f"A{sec_row}:{get_column_letter(_S1_NCOLS)}{sec_row}")
            c = ws.cell(sec_row, 1)
            c.font = _SECTION_FONT
            c.fill = _SECTION_FILL

            # 컬럼 헤더
            hdr_row = sec_row + 1
            _write_header_row(ws, hdr_row, _S1_COLS)

            cur.execute("""
                SELECT p.ticker, p.quantity, p.avg_price,
                       COALESCE(t.name, p.ticker),
                       COALESCE(t.market, ''),
                       COALESCE(t.leverage, 1)
                FROM positions p
                LEFT JOIN tickers t ON t.ticker = p.ticker
                WHERE p.account_id = %s
                ORDER BY p.ticker
            """, (acct_id,))
            positions = cur.fetchall()

            acct_total = 0.0
            for ticker, qty, avg_price, tname, market, leverage in positions:
                qty_f  = float(qty or 0)
                avg_f  = float(avg_price) if avg_price is not None else None

                if ticker == "KRW":
                    price, tname, display_market, display_lev = 1.0, "원화 현금", "-", "-"
                    pnl, pnl_pct = None, None
                elif ticker == "USD":
                    price, tname, display_market, display_lev = usd_krw, "달러 현금", "-", "-"
                    pnl, pnl_pct = None, None
                else:
                    p_data = prices.get(ticker)
                    price  = float(p_data["price"]) if p_data else 0.0
                    display_market = market
                    display_lev    = int(leverage) if leverage else 1
                    if avg_f and avg_f > 0 and price > 0:
                        base = qty_f * usd_krw if is_us_market(market) else qty_f
                        pnl     = (price - avg_f) * base
                        pnl_pct = (price / avg_f - 1) * 100
                    else:
                        pnl, pnl_pct = None, None

                eval_krw = _eval_krw(ticker, qty_f, price, market, usd_krw)
                acct_total += eval_krw

                ws.append([
                    tname, ticker,
                    display_market, display_lev,
                    _f(qty_f, 4),
                    _f(avg_f, 2) if avg_f is not None else "-",
                    _f(price, 2),
                    _f(eval_krw, 0),
                    _f(pnl, 0) if pnl is not None else "-",
                    _f(pnl_pct, 2) if pnl_pct is not None else "-",
                ])

            # 소계
            ws.append(["소계", "", "", "", "", "", "", _f(acct_total, 0), "", ""])
            tot_row = ws.max_row
            ws.cell(tot_row, 1).font = _TOTAL_FONT
            ws.cell(tot_row, 8).font = _TOTAL_FONT
            ws.append([""] * _S1_NCOLS)

        cur.close()

    _auto_width(ws)


# ── Sheet 2: 보유종목 통합 ────────────────────────────────────────────────────

_S2_COLS = ["종목명", "티커", "시장", "레버리지", "수량", "평단", "현재가",
            "평가액(원)", "손익(원)", "손익률(%)", "비중(%)"]


def _build_sheet2(wb, prices: dict, usd_krw: float, now_str: str):
    ws = wb.create_sheet("보유종목 통합")

    ws.cell(1, 1, f"기준: {now_str}").font = _META_FONT
    ws.append([""] * len(_S2_COLS))
    _write_header_row(ws, 3, _S2_COLS)

    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT p.ticker,
                   COALESCE(t.name, p.ticker),
                   COALESCE(t.market, ''),
                   COALESCE(t.leverage, 1),
                   SUM(p.quantity) AS total_qty,
                   CASE WHEN SUM(p.quantity) > 0
                        THEN SUM(p.quantity * COALESCE(p.avg_price, 0)) / SUM(p.quantity)
                        ELSE NULL END AS wav_price
            FROM positions p
            LEFT JOIN tickers t ON t.ticker = p.ticker
            LEFT JOIN accounts a ON p.account_id = a.id
            WHERE (a.is_watch = false OR a.is_watch IS NULL)
            GROUP BY p.ticker, t.name, t.market, t.leverage
            ORDER BY p.ticker
        """)
        rows = cur.fetchall()
        cur.close()

    # 총 평가액 계산
    eval_list  = []
    total_eval = 0.0
    for ticker, tname, market, leverage, qty, avg_price in rows:
        qty_f = float(qty or 0)
        if ticker == "KRW":
            e = qty_f
        elif ticker == "USD":
            e = qty_f * usd_krw
        else:
            p_data = prices.get(ticker)
            price  = float(p_data["price"]) if p_data else 0.0
            e = _eval_krw(ticker, qty_f, price, market, usd_krw)
        eval_list.append(e)
        total_eval += e

    for i, (ticker, tname, market, leverage, qty, avg_price) in enumerate(rows):
        qty_f  = float(qty or 0)
        avg_f  = float(avg_price) if avg_price is not None else None
        eval_krw = eval_list[i]
        weight   = (eval_krw / total_eval * 100) if total_eval > 0 else 0.0

        if ticker == "KRW":
            price, tname, display_market, display_lev = 1.0, "원화 현금", "-", "-"
            pnl, pnl_pct = None, None
        elif ticker == "USD":
            price, tname, display_market, display_lev = usd_krw, "달러 현금", "-", "-"
            pnl, pnl_pct = None, None
        else:
            p_data = prices.get(ticker)
            price  = float(p_data["price"]) if p_data else 0.0
            display_market = market
            display_lev    = int(leverage) if leverage else 1
            if avg_f and avg_f > 0 and price > 0:
                base = qty_f * usd_krw if is_us_market(market) else qty_f
                pnl     = (price - avg_f) * base
                pnl_pct = (price / avg_f - 1) * 100
            else:
                pnl, pnl_pct = None, None

        ws.append([
            tname, ticker,
            display_market, display_lev,
            _f(qty_f, 4),
            _f(avg_f, 2) if avg_f is not None else "-",
            _f(price, 2),
            _f(eval_krw, 0),
            _f(pnl, 0) if pnl is not None else "-",
            _f(pnl_pct, 2) if pnl_pct is not None else "-",
            _f(weight, 2),
        ])

    # 합계
    ws.append(["합계", "", "", "", "", "", "",
               _f(total_eval, 0), "", "", "100.00"])
    tot_row = ws.max_row
    for col in range(1, len(_S2_COLS) + 1):
        ws.cell(tot_row, col).font = _TOTAL_FONT

    _auto_width(ws)


# ── Sheet 3: 실적기록 ─────────────────────────────────────────────────────────

_S3_COLS = [
    "날짜", "총자산(원)", "전일비(원)", "총자산증감(%)",
    "TWR(%)", "NDX100", "NDX100증감(%)",
    "입출금(원)", "입출금사유",
    "Exposure(%)", "현금비중(%)", "x1(%)", "x2(%)", "x3(%)", "USD/KRW",
]


def _build_sheet3(wb):
    ws = wb.create_sheet("실적기록")
    _write_header_row(ws, 1, _S3_COLS)

    db_rows   = load_history()   # ASC
    today_row = load_today_row() # dict or None

    # 표시 순서: 최신(today) → 과거 (DESC)
    all_rows = []
    if today_row:
        all_rows.append(today_row)
    for r in reversed(db_rows):
        all_rows.append({
            "date":           str(r[0]),
            "total_asset":    r[1],
            "twr_asset":      r[2],
            "ndx100":         r[3],
            "cash_flow":      r[4],
            "cash_flow_note": r[5],
            "exposure":       r[6],
            "cash_ratio":     r[7],
            "x1_ratio":       r[8],
            "x2_ratio":       r[9],
            "x3_ratio":       r[10],
            "usd_krw":        r[11],
        })

    # TWR / NDX 기준값 (가장 오래된 DB 행)
    base_twr = float(db_rows[0][2] or 0) if db_rows else None
    base_ndx = float(db_rows[0][3] or 0) if db_rows else None

    for i, row in enumerate(all_rows):
        total    = float(row["total_asset"] or 0)
        twr_asset = float(row.get("twr_asset") or 0)
        ndx      = float(row.get("ndx100") or 0)
        cf       = int(row.get("cash_flow") or 0)

        # 전일비: 바로 아래 행(전일)과 비교
        if i < len(all_rows) - 1:
            prev_total = float(all_rows[i + 1]["total_asset"] or 0)
            day_diff   = _f(total - prev_total, 0)
            day_pct    = _f((total / prev_total - 1) * 100, 2) if prev_total else None
        else:
            day_diff = day_pct = None

        twr_pct = _f((twr_asset / base_twr - 1) * 100, 2) if base_twr else None
        ndx_pct = _f((ndx / base_ndx - 1) * 100, 2) if base_ndx else None

        ws.append([
            row["date"],
            _f(total, 0),
            day_diff if day_diff is not None else "-",
            day_pct  if day_pct  is not None else "-",
            twr_pct  if twr_pct  is not None else "-",
            _f(ndx, 2),
            ndx_pct  if ndx_pct  is not None else "-",
            cf if cf else "-",
            row.get("cash_flow_note") or "",
            _f(float(row.get("exposure")   or 0) * 100, 1),
            _f(float(row.get("cash_ratio") or 0) * 100, 1),
            _f(float(row.get("x1_ratio")   or 0) * 100, 1),
            _f(float(row.get("x2_ratio")   or 0) * 100, 1),
            _f(float(row.get("x3_ratio")   or 0) * 100, 1),
            _f(float(row.get("usd_krw")    or 0), 2),
        ])

    _auto_width(ws)


# ── 메인 ─────────────────────────────────────────────────────────────────────

def build_export_xlsx() -> bytes:
    prices   = get_all_prices()
    usd_krw  = _usd_krw_from_redis(prices)
    now_str  = datetime.datetime.now(KST).strftime("%Y-%m-%d %H:%M KST")

    wb = openpyxl.Workbook()
    _build_sheet1(wb, prices, usd_krw, now_str)
    _build_sheet2(wb, prices, usd_krw, now_str)
    _build_sheet3(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()